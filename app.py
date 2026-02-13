# import streamlit as st
# import pandas as pd

# st.set_page_config("Min Day Finder", layout="centered")
# st.title("📅 Campaign First Live Date Finder")


# # ======================
# # Load Excel & Fix Blank Rows
# # ======================

# def load_excel_with_header(file):
#     raw = pd.read_excel(file, header=None)

#     header_row = None

#     for i in range(len(raw)):
#         row_values = raw.iloc[i].astype(str).str.lower().tolist()

#         if "day" in row_values and "campaign" in row_values:
#             header_row = i
#             break

#     if header_row is None:
#         st.error("Header row with 'Day' and 'Campaign' not found")
#         st.stop()

#     return pd.read_excel(file, header=header_row)


# # ======================
# # Upload
# # ======================

# uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])

# if not uploaded_file:
#     st.stop()

# df = load_excel_with_header(uploaded_file)

# # ======================
# # User Input
# # ======================

# unique_key = st.text_input("Enter Unique Key", value="1ur-456492l")

# if not unique_key:
#     st.stop()

# # ======================
# # Convert Day
# # ======================

# df["Day"] = pd.to_datetime(df["Day"], dayfirst=True, errors="coerce")

# # ======================
# # Filter
# # ======================

# filtered = df[
#     df["Campaign"].astype(str).str.contains(unique_key, case=False, na=False)
# ]

# # ======================
# # Output
# # ======================

# if filtered.empty:
#     st.error("No data found for this Unique Key")
# else:
#     min_day = filtered["Day"].min()
#     st.success(f"✅ First Live Date: {min_day.strftime('%d-%b-%Y')}")
























import streamlit as st
import pandas as pd
import re

st.set_page_config("SA Report Automation", layout="wide")
st.title("📊 SA Monthly Report")


# ==========================
# File type detection
# ==========================

PLATFORM_MAP = {
    "media_plan": ["media", "plan", "qt"],
    "craft": ["craft"],
    "google": ["google"],
    "social": ["social", "facebook", "instagram"],
    "dv360": ["dv360"],
    "sizmek": ["sizmek", "flashtalking"],
    "ecom": ["ecom", "amazon", "flipkart"],
}


def normalize(x):
    return re.sub(r"[^a-z0-9]", "", str(x).lower())


def detect_file_type(filename):
    name = normalize(filename)

    for platform, keys in PLATFORM_MAP.items():
        for k in keys:
            if k in name:
                return platform

    return "unknown"


# ==========================
# Upload multiple files
# ==========================

uploaded_files = st.file_uploader(
    "Upload Media Plan, CRAFT & Publisher Reports",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=True
)

if not uploaded_files:
    st.stop()


# ==========================
# Classify files
# ==========================

files_by_type = {}

for f in uploaded_files:
    ftype = detect_file_type(f.name)
    files_by_type.setdefault(ftype, []).append(f)


# ==========================
# Show detection (clean)
# ==========================

st.subheader("📁 Detected Inputs")

for k, v in files_by_type.items():
    st.write(f"**{k.upper()}**")
    for f in v:
        st.write("•", f.name)


# ==========================
# Mandatory check
# ==========================

if "media_plan" not in files_by_type or "craft" not in files_by_type:
    st.error("Media Plan & CRAFT files are mandatory")
    st.stop()

st.success("✅ Files successfully classified — ready to build SA Report")



# Phase 2

# =====================================================
# PHASE 2 — MEDIA PLAN TAB EXTRACTION
# =====================================================

media_file = files_by_type["media_plan"][0]
media_xl = pd.ExcelFile(media_file)

def normalize(x):
    return re.sub(r"[^a-z0-9]", "", str(x).lower())


def is_data_sheet(name):
    n = normalize(name)
    return "checklist" not in n and "summary" not in n


# ad_type_sheets = [s for s in media_xl.sheet_names if is_data_sheet(s)]
ad_type_sheets = [
    s for s in media_xl.sheet_names
    if is_data_sheet(s) and normalize(s) != "estoresearch"
]

st.subheader("📊 Media Plan Ad Types Detected")

for s in ad_type_sheets:
    st.write("•", s)

st.success(f"✅ {len(ad_type_sheets)} Ad Type tabs ready for processing")




# Detect Brand name
def detect_brand(filename):
    name = normalize(filename)

    if "watch" in name:
        return "Watch"
    if "mobile" in name or "phone" in name:
        return "Mobile"
    if "tv" in name:
        return "TV"

    return "Samsung"
    

brand_name = detect_brand(media_file.name)
# def extract_qt(filename):
#     match = re.search(r"qt\d+", filename.lower())
#     if match:
#         return match.group(0).upper()
#     return "QTXXXXXXX"
# qt_number = extract_qt(media_file.name)



# Phase3

# =====================================================
# PHASE 3 — CAMPAIGN DATE EXTRACTION FROM CHECKLIST
# =====================================================

campaign_start = None
campaign_end = None

for sheet in media_xl.sheet_names:

    if "checklist" in normalize(sheet):

        checklist = media_xl.parse(sheet, header=None)

        for _, row in checklist.iterrows():

            text = " ".join(row.astype(str).str.lower())
            dates = pd.to_datetime(row, errors="coerce", dayfirst=True).dropna()

            if not dates.empty:

                if "from" in text and campaign_start is None:
                    campaign_start = dates.iloc[0]

                if "till" in text and campaign_end is None:
                    campaign_end = dates.iloc[0]


if not campaign_start or not campaign_end:
    st.error("Could not detect campaign start/end dates from Checklist")
    st.stop()


st.subheader("📆 Campaign Duration Detected")

st.write("Start Date:", campaign_start.strftime("%d-%b-%Y"))
st.write("End Date:", campaign_end.strftime("%d-%b-%Y"))

campaign_days = (campaign_end - campaign_start).days + 1

st.success(f"✅ Campaign Days: {campaign_days}")



# Phase 4

# # =====================================================
# # PHASE 4 — MEDIA PLAN LINE ITEM PARSER
# # =====================================================

# def find_column(df, keywords):
#     for col in df.columns:
#         col_norm = normalize(col)
#         for k in keywords:
#             if normalize(k) in col_norm:
#                 return col
#     return None


# media_plan_items = {}

# for sheet in ad_type_sheets:

#     df = media_xl.parse(sheet, header=None)

#     # --- detect header row ---
#     header_row = None
#     for i in range(min(25, len(df))):
#         row = df.iloc[i].astype(str).str.lower()
#         if "unique" in " ".join(row) and "buy" in " ".join(row):
#             header_row = i
#             break

#     if header_row is None:
#         continue

#     df = media_xl.parse(sheet, header=header_row)
#     df = df.dropna(how="all")

#     # --- column mapping ---
#     col_map = {
#         "publisher": find_column(df, ["publisher", "site"]),
#         "unique_key": find_column(df, ["unique", "key"]),
#         "objective": find_column(df, ["objective"]),
#         "property": find_column(df, ["property", "placement"]),
#         "ad_unit": find_column(df, ["ad unit"]),
#         "buy_type": find_column(df, ["buy"]),
#         "est_clicks": find_column(df, ["est", "click"]),
#         "est_imps": find_column(df, ["est", "imp"]),
#         "est_views": find_column(df, ["est", "view"]),
#         "est_eng": find_column(df, ["est", "eng"]),
#         "est_leads": find_column(df, ["est", "lead"]),
#     }

#     rows = []

#     for _, r in df.iterrows():

#         uk = r.get(col_map["unique_key"])
#         if pd.isna(uk):
#             continue

#         buy = str(r.get(col_map["buy_type"])).lower()

#         # --- planned v1 selection based on Buy Type ---
#         planned_v1 = None

#         if "cpc" in buy:
#             planned_v1 = r.get(col_map["est_clicks"])
#         elif "cpm" in buy or "fixed" in buy or "flat" in buy or "cph" in buy:
#             planned_v1 = r.get(col_map["est_imps"])
#         elif "cpv" in buy:
#             planned_v1 = r.get(col_map["est_views"])
#         elif "cpe" in buy:
#             planned_v1 = r.get(col_map["est_eng"])
#         elif "cpl" in buy:
#             planned_v1 = r.get(col_map["est_leads"])

#         rows.append({
#             "Genres": sheet,
#             "Site": r.get(col_map["publisher"]),
#             "Unique Key": uk,
#             "Objective/Targeting": r.get(col_map["objective"]),
#             "Property/Inventory": r.get(col_map["property"], "-"),
#             "Ad Unit": r.get(col_map["ad_unit"]),
#             "Cost Format": buy.upper(),
#             "Planned Delivery v1": planned_v1,
#         })

#     media_plan_items[sheet] = pd.DataFrame(rows)


# # =====================================================
# # DISPLAY SUMMARY (NO DATA PREVIEW)
# # =====================================================

# st.subheader("📌 Media Plan Line Items Parsed")

# for sheet, df in media_plan_items.items():
#     st.write(f"• {sheet}: {len(df)} line items")

# st.success("✅ Media Plan structure successfully parsed")








# =====================================================
# PHASE 4 — MEDIA PLAN LINE ITEM PARSER (FIXED)
# =====================================================

def find_column_strict(df, must_have, optional=None):
    """
    must_have: list of mandatory keywords (eg ["click"])
    optional: list of optional keywords (eg ["est", "planned"])
    """
    for col in df.columns:
        col_norm = normalize(col)

        if all(k in col_norm for k in must_have):
            if optional:
                if any(o in col_norm for o in optional):
                    return col
            else:
                return col
    return None


media_plan_items = {}

for sheet in ad_type_sheets:

    raw = media_xl.parse(sheet, header=None)

    # ---- detect header row ----
    header_row = None
    for i in range(min(25, len(raw))):
        row = raw.iloc[i].astype(str).str.lower()
        if "unique" in " ".join(row) and "buy" in " ".join(row):
            header_row = i
            break

    if header_row is None:
        continue

    df = media_xl.parse(sheet, header=header_row).dropna(how="all")

    # ---- column mapping (STRICT) ----
    col_map = {
        "publisher": find_column_strict(df, ["publisher"]) or find_column_strict(df, ["site"]),
        "unique_key": find_column_strict(df, ["unique", "key"]),
        "objective": find_column_strict(df, ["objective"]),
        "property": find_column_strict(df, ["property"]) or find_column_strict(df, ["placement"]),
        "ad_unit": find_column_strict(df, ["ad", "unit"]),
        "buy_type": find_column_strict(df, ["buy"]),

        # Planned metrics (STRICT)
        "est_clicks": find_column_strict(df, ["click"], ["est", "planned"]),
        "est_imps": find_column_strict(df, ["impression"], ["est", "planned"]),
        "est_views": find_column_strict(df, ["view"], ["est", "planned"]),
        "est_eng": find_column_strict(df, ["engagement"], ["est", "planned"]),
        "est_leads": find_column_strict(df, ["lead"], ["est", "planned"]),
    }

    rows = []

    for _, r in df.iterrows():

        uk = r.get(col_map["unique_key"])
        if pd.isna(uk):
            continue

        buy = str(r.get(col_map["buy_type"])).lower()
        planned_v1 = None

        # ---- CORRECT BUY TYPE → METRIC MAPPING ----
        if "cpc" in buy:
            planned_v1 = r.get(col_map["est_clicks"])
        elif any(x in buy for x in ["cpm", "fixed", "flat", "cph"]):
            planned_v1 = r.get(col_map["est_imps"])
        elif "cpv" in buy:
            planned_v1 = r.get(col_map["est_views"])
        elif "cpe" in buy:
            planned_v1 = r.get(col_map["est_eng"])
        elif "cpl" in buy:
            planned_v1 = r.get(col_map["est_leads"])

        rows.append({
            "Genres": sheet,
            "Site": r.get(col_map["publisher"]),
            "Unique Key": uk,
            "Objective/Targeting": r.get(col_map["objective"]),
            "Property/Inventory": r.get(col_map["property"], "-"),
            "Ad Unit": r.get(col_map["ad_unit"]),
            "Cost Format": buy.upper(),
            "Planned Delivery v1": planned_v1,
        })

    media_plan_items[sheet] = pd.DataFrame(rows)


st.success("✅ Phase 4 FIXED — Planned Delivery v1 now matches Buy Type correctly")













# Phase 5

# =====================================================
# PHASE 5 — CRAFT REPORT INTEGRATION
# =====================================================

craft_file = files_by_type["craft"][0]
craft_raw = pd.read_excel(craft_file, header=None)

# --- detect header row ---
craft_header = None
for i in range(min(25, len(craft_raw))):
    row = craft_raw.iloc[i].astype(str).str.lower()
    if "unique" in " ".join(row) and "planned" in " ".join(row):
        craft_header = i
        break

if craft_header is None:
    st.error("CRAFT header not detected")
    st.stop()

craft_df = pd.read_excel(craft_file, header=craft_header)
craft_df.columns = [str(c) for c in craft_df.columns]


def find_craft_col(keyword):
    for col in craft_df.columns:
        if keyword in normalize(col):
            return col
    return None

def find_reported(keyword):
    for col in craft_df.columns:
        name = normalize(col)
        if "reported" in name and keyword in name:
            return col
    return None


# craft_cols = {
#     "qt": find_craft_col("qt"),
#     "channel": find_craft_col("channel"),
#     "uk": find_craft_col("unique"),
#     "p_click": find_craft_col("plannedclick"),
#     "p_imp": find_craft_col("plannedimp"),
#     "p_view": find_craft_col("plannedvideo"),
#     "p_eng": find_craft_col("plannedeng"),
#     "r_click": find_craft_col("click"),
#     "r_imp": find_craft_col("imp"),
#     "r_view": find_craft_col("video"),
#     "r_eng": find_craft_col("eng"),
# }


def find_reported_only(keyword):
    for col in craft_df.columns:
        name = normalize(col)

        if keyword in name and "planned" not in name:
            return col
    return None



craft_cols = {
    "qt": find_craft_col("qt"),
    "channel": find_craft_col("channel"),
    "uk": find_craft_col("unique"),

    # PLANNED (same as before)
    "p_click": find_craft_col("plannedclick"),
    "p_imp": find_craft_col("plannedimp"),
    "p_view": find_craft_col("plannedvideo"),
    "p_eng": find_craft_col("plannedeng"),

    # ✅ REPORTED — NON PLANNED ONLY
    "r_click": find_reported_only("click"),
    "r_imp": find_reported_only("impression"),
    "r_view": find_reported_only("video"),
    "r_eng": find_reported_only("engagement"),
}


# def get_craft_values(ad_type, uk, buy_type):

#     df = craft_df.copy()

#     # --- normal campaigns ---
#     if craft_cols["channel"]:
#         df = df[df[craft_cols["channel"]].astype(str).str.contains(ad_type, case=False, na=False)]

#     df = df[df[craft_cols["uk"]].astype(str) == str(uk)]

#     if df.empty:
#         return None, None

#     buy = buy_type.lower()

#     if "cpc" in buy:
#         return df[craft_cols["p_click"]].sum(), df[craft_cols["r_click"]].sum()

#     if "cpm" in buy or "fixed" in buy or "flat" in buy or "cph" in buy:
#         return df[craft_cols["p_imp"]].sum(), df[craft_cols["r_imp"]].sum()

#     if "cpv" in buy:
#         return df[craft_cols["p_view"]].sum(), df[craft_cols["r_view"]].sum()

#     if "cpe" in buy:
#         return df[craft_cols["p_eng"]].sum(), df[craft_cols["r_eng"]].sum()

#     return None, None





# def get_craft_values(ad_type, uk, buy_type):

#     df = craft_df.copy()

#     if craft_cols["channel"]:
#         df = df[df[craft_cols["channel"]].astype(str).str.contains(ad_type, case=False, na=False)]

#     df = df[df[craft_cols["uk"]].astype(str) == str(uk)]

#     if df.empty:
#         return None, None

#     buy = buy_type.lower()

#     def safe_sum(col):
#         if col and col in df.columns:
#             return pd.to_numeric(df[col], errors="coerce").sum()
#         return None

#     # ---- CPC ----
#     if "cpc" in buy:
#         return safe_sum(craft_cols["p_click"]), safe_sum(craft_cols["r_click"])

#     # ---- CPV ----
#     if "cpv" in buy:
#         return safe_sum(craft_cols["p_view"]), safe_sum(craft_cols["r_view"])

#     # ---- CPE ----
#     if "cpe" in buy:
#         return safe_sum(craft_cols["p_eng"]), safe_sum(craft_cols["r_eng"])

#     # ---- CPM / FIXED / FLAT / CPH ----
#     return safe_sum(craft_cols["p_imp"]), safe_sum(craft_cols["r_imp"])




def get_craft_values(ad_type, uk, buy_type):

    df = craft_df.copy()

    # ---------- TRY NORMAL FILTER (CHANNEL + UNIQUE) ----------
    if craft_cols["channel"]:
        df_filtered = df[
            df[craft_cols["channel"]]
            .astype(str)
            .str.contains(ad_type, case=False, na=False)
        ]
    else:
        df_filtered = df

    df_filtered = df_filtered[
        df_filtered[craft_cols["uk"]].astype(str) == str(uk)
    ]

    # ---------- FALLBACK: UNIQUE KEY ONLY ----------
    if df_filtered.empty:
        df_filtered = df[
            df[craft_cols["uk"]].astype(str) == str(uk)
        ]

    if df_filtered.empty:
        return None, None

    buy = buy_type.lower()

    if "cpc" in buy:
        return (
            df_filtered[craft_cols["p_click"]].sum(),
            df_filtered[craft_cols["r_click"]].sum()
        )

    if "cpm" in buy or "fixed" in buy or "flat" in buy or "cph" in buy:
        return (
            df_filtered[craft_cols["p_imp"]].sum(),
            df_filtered[craft_cols["r_imp"]].sum()
        )

    if "cpv" in buy:
        return (
            df_filtered[craft_cols["p_view"]].sum(),
            df_filtered[craft_cols["r_view"]].sum()
        )

    if "cpe" in buy:
        return (
            df_filtered[craft_cols["p_eng"]].sum(),
            df_filtered[craft_cols["r_eng"]].sum()
        )

    return None, None



# =====================================================
# ATTACH CRAFT VALUES TO MEDIA PLAN STRUCTURE
# =====================================================

for ad_type, df in media_plan_items.items():

    craft_planned = []
    craft_reported = []

    for _, row in df.iterrows():

        cp, cr = get_craft_values(
            ad_type,
            row["Unique Key"],
            row["Cost Format"]
        )

        craft_planned.append(cp)
        craft_reported.append(cr)

    df["CRAFT Planned Delivery"] = craft_planned
    df["CRAFT Reported Delivery"] = craft_reported


st.success("✅ CRAFT planned & reported values mapped to all ad types")




# =====================================================
# PHASE 6 — VERIFIED PUBLISHER EXTRACTION (FINAL WORKING)
# =====================================================



















# # =====================================================
# # PHASE 6 — VERIFIED PUBLISHER EXTRACTION (FINAL WORKING)
# # =====================================================


# def load_publisher(file):
#     raw = pd.read_excel(file, header=None)

#     for i in range(min(30, len(raw))):
#         row = raw.iloc[i].astype(str).str.lower()
#         joined = " ".join(row)

#         if any(x in joined for x in ["campaign", "placement", "insertion", "unique"]):
#             if "date" in joined or "day" in joined:
#                 return pd.read_excel(file, header=i)

#     return pd.read_excel(file)




# # AMAZON FILE DETECTION (must be defined before use)

# def load_ecom_publisher(file):
#     xl = pd.ExcelFile(file)
#     dfs = []

#     for sheet in xl.sheet_names:
#         df = xl.parse(sheet)

#         # Only the Amazon tab contains Unique Key
#         if any("unique" in normalize(c) for c in df.columns):
#             dfs.append(df)

#     return dfs

# # ================= Load publisher files =================

# # publisher_data = {}

# # for platform, files in files_by_type.items():
# #     if platform in ["media_plan", "craft"]:
# #         continue

# #     publisher_data[platform] = [load_publisher(f) for f in files]

# publisher_data = {}

# for platform, files in files_by_type.items():
#     if platform in ["media_plan", "craft"]:
#         continue

#     publisher_data[platform] = []

#     for f in files:
#         if platform == "ecom":
#             # Amazon file has multiple tabs → load all relevant sheets
#             publisher_data[platform].extend(load_ecom_publisher(f))
#         else:
#             # All other platforms remain unchanged
#             publisher_data[platform].append(load_publisher(f))


# # ================= Platform routing =================

# def route_platform(site):
#     s = normalize(site)

#     if "dv360" in s:
#         return "dv360"
#     if "google" in s or "youtube" in s:
#         return "google"
#     if "facebook" in s or "instagram" in s:
#         return "social"
#     if "inshorts" in s or "91mobiles" in s or "sizmek" in s or "flashtalking" in s:
#         return "sizmek"
#     if "amazon" in s or "flipkart" in s:
#         return "ecom"

#     return None


# # ================= Smart column finder =================

# def find_col(df, keywords):
#     for c in df.columns:
#         name = normalize(c)
#         for k in keywords:
#             if k in name:
#                 return c
#     return None


# # ================= Generic extractor =================

# def extract_generic(df, uk, key_keys, date_keys, metric_keys):

#     key_col = find_col(df, key_keys)
#     date_col = find_col(df, date_keys)
#     metric_col = find_col(df, metric_keys)

#     if not key_col or not metric_col:
#         return None, None

#     rows = df[df[key_col].astype(str).apply(
#         lambda x: normalize(uk) in normalize(x)
#     )]

#     if rows.empty:
#         return None, None

#     value = pd.to_numeric(rows[metric_col], errors="coerce").sum()

#     live = (
#         pd.to_datetime(rows[date_col], errors="coerce").min()
#         if date_col else None
#     )

#     return value, live


# # ================= Google special =================

# def extract_google(df, uk, buy_type):

#     # -------- identify columns --------
#     campaign_col = None
#     day_col = None

#     for c in df.columns:
#         n = normalize(c)
#         if "campaign" in n:
#             campaign_col = c
#         if "day" in n or "date" in n:
#             day_col = c

#     if not campaign_col or not day_col:
#         return None, None

#     # -------- filter rows by unique key --------
#     rows = df[df[campaign_col].astype(str).str.contains(uk, na=False)]

#     if rows.empty:
#         return None, None

#     buy = buy_type.lower()

#     # -------- pick metric column --------
#     if "cpc" in buy:
#         metric_col = next(c for c in df.columns if "click" in normalize(c))
#     elif "cpv" in buy:
#         metric_col = next(c for c in df.columns if "view" in normalize(c))
#     else:
#         metric_col = next(c for c in df.columns if "impr" in normalize(c))

#     # -------- SAFE numeric conversion --------
#     metric_vals = (
#         rows[metric_col]
#         .astype(str)
#         .str.replace(",", "", regex=False)
#         .str.replace("--", "", regex=False)
#         .str.strip()
#     )

#     value = pd.to_numeric(metric_vals, errors="coerce").sum()

#     # -------- earliest live date --------
#     live = (
#         pd.to_datetime(
#             rows[day_col],
#             errors="coerce",
#             dayfirst=True
#         )
#         .min()
#     )

#     return value if value > 0 else None, live


# # AMAZON FILE DETECTION

# def load_ecom_publisher(file):
#     xl = pd.ExcelFile(file)
#     dfs = []

#     for sheet in xl.sheet_names:
#         df = xl.parse(sheet)

#         # Only the Amazon tab contains Unique Key
#         if any("unique" in normalize(c) for c in df.columns):
#             dfs.append(df)

#     return dfs




# # ================= Attach Actual + Live =================

# for ad_type, df in media_plan_items.items():

#     actuals, lives = [], []

#     for _, r in df.iterrows():

#         platform = route_platform(r["Site"])
#         uk = r["Unique Key"]
#         buy = r["Cost Format"]

#         total = 0
#         dates = []

#         for pub_df in publisher_data.get(platform, []):

#             if platform == "google":
#                 val, live = extract_google(pub_df, uk, buy)

#             elif platform == "social":
#                 val, live = extract_generic(
#                     pub_df, uk,
#                     key_keys=["unique"],
#                     date_keys=["final", "date", "day"],
#                     metric_keys=["click", "view", "engage"]
#                 )

#             elif platform == "dv360":
#                 val, live = extract_generic(
#                     pub_df, uk,
#                     key_keys=["insertion"],
#                     date_keys=["date"],
#                     metric_keys=["impression", "view"]
#                 )

#             elif platform == "sizmek":
#                 val, live = extract_generic(
#                     pub_df, uk,
#                     key_keys=["placement"],
#                     date_keys=["day", "date"],
#                     metric_keys=["impression", "click"]
#                 )

#             elif platform == "ecom":
#                 val, live = extract_generic(
#                     pub_df, uk,
#                     key_keys=["unique"],
#                     date_keys=["date"],
#                     metric_keys=["click"]
#                 )

#             else:
#                 continue

#             if val is not None:
#                 total += val
#             if live is not None:
#                 dates.append(live)

#         actuals.append(total if total > 0 else None)
#         lives.append(min(dates) if dates else None)

#     df["Actual Delivered Reporting SA"] = actuals
#     df["Live Date"] = lives


# st.success("✅ Phase 6 COMPLETE — All publisher data mapped correctly")




































# ================================================s=====
# PHASE 6 — VERIFIED PUBLISHER EXTRACTION (FINAL WORKING)
# =====================================================


def load_publisher(file):
    raw = pd.read_excel(file, header=None)

    for i in range(min(30, len(raw))):
        row = raw.iloc[i].astype(str).str.lower()
        joined = " ".join(row)

        if any(x in joined for x in ["campaign", "placement", "insertion", "unique"]):
            if "date" in joined or "day" in joined:
                return pd.read_excel(file, header=i)

    return pd.read_excel(file)




# AMAZON FILE DETECTION (must be defined before use)

def load_ecom_publisher(file):
    xl = pd.ExcelFile(file)
    dfs = []

    for sheet in xl.sheet_names:
        df = xl.parse(sheet)

        # Only the Amazon tab contains Unique Key
        if any("unique" in normalize(c) for c in df.columns):
            dfs.append(df)

    return dfs

# ================= Load publisher files =================

# publisher_data = {}

# for platform, files in files_by_type.items():
#     if platform in ["media_plan", "craft"]:
#         continue

#     publisher_data[platform] = [load_publisher(f) for f in files]

publisher_data = {}

for platform, files in files_by_type.items():
    if platform in ["media_plan", "craft"]:
        continue

    publisher_data[platform] = []

    for f in files:
        if platform == "ecom":
            # Amazon file has multiple tabs → load all relevant sheets
            publisher_data[platform].extend(load_ecom_publisher(f))
        else:
            publisher_data[platform] = [load_publisher(f)]



st.write("PLATFORMS LOADED")



# ================= Platform routing =================

def route_platform(site):
    s = normalize(site)

    if "dv360" in s:
        return "dv360"
    if "google" in s or "youtube" in s:
        return "google"
    if "facebook" in s or "instagram" in s:
        return "social"

    if "91mobiles" in s or "inshorts" in s or "sizmek" in s or "flashtalking" in s:
        return "sizmek"

    if "amazon" in s or "flipkart" in s:
        return "ecom"

    return None


# ================= Smart column finder =================

def find_col(df, keywords):
    for c in df.columns:
        name = normalize(c)
        for k in keywords:
            if k in name:
                return c
    return None


# ================= Generic extractor =================

def extract_generic(df, uk, key_keys, date_keys, metric_keys):

    key_col = find_col(df, key_keys)
    date_col = find_col(df, date_keys)
    metric_col = find_col(df, metric_keys)

    if not key_col or not metric_col:
        return None, None

    rows = df[df[key_col].astype(str).apply(
        lambda x: normalize(uk) in normalize(x)
    )]

    if rows.empty:
        return None, None

    value = pd.to_numeric(rows[metric_col], errors="coerce").sum()

    live = (
        pd.to_datetime(rows[date_col], errors="coerce").min()
        if date_col else None
    )

    return value, live






# ================= Google special =================

def extract_google(df, uk, buy_type):

    # -------- identify columns --------
    campaign_col = None
    day_col = None

    for c in df.columns:
        n = normalize(c)
        if "campaign" in n:
            campaign_col = c
        if "day" in n or "date" in n:
            day_col = c

    if not campaign_col or not day_col:
        return None, None

    # -------- filter rows by unique key --------
    rows = df[df[campaign_col].astype(str).str.contains(uk, na=False)]

    if rows.empty:
        return None, None

    buy = buy_type.lower()

    # -------- pick metric column --------
    if "cpc" in buy:
        metric_col = next(c for c in df.columns if "click" in normalize(c))
    elif "cpv" in buy:
        metric_col = next(c for c in df.columns if "view" in normalize(c))
    else:
        metric_col = next(c for c in df.columns if "impr" in normalize(c))

    # -------- SAFE numeric conversion --------
    metric_vals = (
        rows[metric_col]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("--", "", regex=False)
        .str.strip()
    )

    value = pd.to_numeric(metric_vals, errors="coerce").sum()

    # -------- earliest live date --------
    live = (
        pd.to_datetime(
            rows[day_col],
            errors="coerce",
            dayfirst=True
        )
        .min()
    )

    return value if value > 0 else None, live




# Video
def extract_video(df, uk, buy_type):

    KEY_COL = find_col(df, ["insertion"])
    DATE_COL = find_col(df, ["date"])

    if not KEY_COL or not DATE_COL:
        return None, None

    rows = df[df[KEY_COL].astype(str).str.contains(uk, na=False)]

    if rows.empty:
        return None, None

    buy = buy_type.lower()

    if "cpv" in buy:
        metric_col = "TrueView: Views"
    elif "cpc" in buy:
        metric_col = find_col(df, ["click"])
    else:
        metric_col = find_col(df, ["impression", "impr"])

    # 🔥 METRIC MAY FAIL — DATE MUST NOT
    if metric_col:
        value = (
            rows[metric_col]
            .astype(str)
            .str.replace(",", "", regex=False)
            .pipe(pd.to_numeric, errors="coerce")
            .sum()
        )
    else:
        value = None

    live = pd.to_datetime(rows[DATE_COL], errors="coerce").min()

    return value if value and value > 0 else None, live






# Social
# def extract_social(df, uk, buy_type):

#     KEY_COL = "Unique Key"
#     DATE_COL = "Final Day Date"
#     AD_ID_COL = "Ad ID"   # IMPORTANT

#     METRIC_MAP = {
#         "cpc": "Clicks (all)",
#         "cpv": "ThruPlays",
#         "cpe": "Post engagements",
#         "cpl": "Meta Leads",
#         "cpm": "Impressions"
#     }

#     if KEY_COL not in df.columns or DATE_COL not in df.columns:
#         return None, None

#     rows = df[df[KEY_COL].astype(str).str.contains(uk, na=False)]

#     if rows.empty:
#         return None, None

#     # ✅ REMOVE META TOTAL ROWS
#     if AD_ID_COL in df.columns:
#         rows = rows[rows[AD_ID_COL].notna()]

#     buy = buy_type.lower()

#     metric_col = None
#     for k, v in METRIC_MAP.items():
#         if k in buy:
#             metric_col = v
#             break

#     if metric_col is None or metric_col not in df.columns:
#         return None, None

#     value = (
#         rows[metric_col]
#         .astype(str)
#         .str.replace(",", "", regex=False)
#         .str.strip()
#         .pipe(pd.to_numeric, errors="coerce")
#         .sum()
#     )

#     live = pd.to_datetime(rows[DATE_COL], errors="coerce").min()

#     return value if value > 0 else None, live



def extract_social(df, uk, buy_type):

    KEY_COL = "Unique Key"
    DATE_COL = "Final Day Date"

    METRIC_MAP = {
        "cpc": "Clicks (all)",
        "cpv": "ThruPlays",
        "cpe": "Post engagements",
        "cpl": "Meta Leads",
        "cpm": "Impressions"
    }

    # 🔥 remove Meta TOTAL row automatically
    df = df[df[KEY_COL].notna()]

    rows = df[df[KEY_COL].astype(str).str.contains(uk, na=False)]

    if rows.empty:
        return None, None

    buy = buy_type.lower()
    metric_col = next(v for k, v in METRIC_MAP.items() if k in buy)

    value = (
        rows[metric_col]
        .astype(str)
        .str.replace(",", "", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
        .sum()
    )

    live = pd.to_datetime(rows[DATE_COL], errors="coerce").min()

    return value if value > 0 else None, live






# AMAZON FILE DETECTION

def load_ecom_publisher(file):
    xl = pd.ExcelFile(file)
    dfs = []

    for sheet in xl.sheet_names:
        df = xl.parse(sheet)

        # Only the Amazon tab contains Unique Key
        if any("unique" in normalize(c) for c in df.columns):
            dfs.append(df)

    return dfs




# ================= Attach Actual + Live =================

for ad_type, df in media_plan_items.items():

    actuals, lives = [], []

    for _, r in df.iterrows():

        platform = route_platform(r["Site"])
        uk = r["Unique Key"]
        buy = r["Cost Format"]

        total = 0
        dates = []

        for pub_df in publisher_data.get(platform, []):

            if platform == "google":
                val, live = extract_google(pub_df, uk, buy)

            elif platform == "social":
                val, live = extract_social(pub_df, uk, buy)


            elif platform == "dv360":
                 val, live = extract_video(pub_df, uk, buy)
            

            # elif platform == "sizmek":
            #     val, live = extract_generic(
            #         pub_df, uk,
            #         key_keys=["placement"],
            #         date_keys=["day", "date"],
            #         metric_keys=["impression", "click"]
            #     )
            elif platform == "sizmek":

                KEY_COL = find_col(pub_df, ["placement"])
                DATE_COL = find_col(pub_df, ["day"])

                if KEY_COL and DATE_COL:

                    rows = pub_df[
                        pub_df[KEY_COL].astype(str).str.contains(uk, case=False, na=False)
                    ]

                    if not rows.empty:

                        if "cpc" in buy.lower():
                            METRIC_COL = find_col(pub_df, ["click"])
                        else:
                            METRIC_COL = find_col(pub_df, ["impression"])

                        if METRIC_COL:

                            val = (
                                rows[METRIC_COL]
                                .astype(str)
                                .str.replace(",", "", regex=False)
                                .pipe(pd.to_numeric, errors="coerce")
                                .sum()
                            )

                            live = pd.to_datetime(
                                rows[DATE_COL],
                                errors="coerce",
                                dayfirst=True
                            ).min()
                        else:
                            val, live = None, None
                    else:
                        val, live = None, None
                else:
                    val, live = None, None




            elif platform == "ecom":
                val, live = extract_generic(
                    pub_df, uk,
                    key_keys=["unique"],
                    date_keys=["date"],
                    metric_keys=["click"]
                )



            else:
                continue

            if val is not None:
                total += val
            if live is not None:
                dates.append(live)

        actuals.append(total if total > 0 else None)
        lives.append(min(dates) if dates else None)

    df["Actual Delivered Reporting SA"] = actuals
    df["Live Date"] = lives


st.success("✅ Phase 6 COMPLETE — All publisher data mapped correctly")


































































































# phase 7
# =====================================================
# PHASE 7 — SAFE SA CALCULATIONS (AUTO COLUMN DETECTION)
# =====================================================

import pandas as pd
import re


def normalize(x):
    return re.sub(r"[^a-z0-9]", "", str(x).lower())


def find_col(df, keyword):
    for c in df.columns:
        if keyword in normalize(c):
            return c
    return None


def safe_div(a, b):
    if pd.isna(a) or pd.isna(b) or b == 0:
        return None
    return a / b


def safe_diff_div(a, b, base):
    if pd.isna(a) or pd.isna(b) or pd.isna(base) or base == 0:
        return None
    return (a - b) / base


for ad_type, df in media_plan_items.items():

    col_actual = find_col(df, "actualdelivered")
    col_planned = find_col(df, "planneddeliveryv1")
    col_craft_plan = find_col(df, "craftplanneddelivery")
    col_craft_rep = find_col(df, "craftreporteddelivery")

    if not all([col_actual, col_planned, col_craft_plan, col_craft_rep]):
        st.warning(f"⚠ Missing columns in {ad_type} — skipping calculations")
        continue

    actual = df[col_actual]
    planned_v1 = df[col_planned]
    craft_planned = df[col_craft_plan]
    craft_reported = df[col_craft_rep]

    df["% v1 Delivery"] = [
        safe_div(a, p) for a, p in zip(actual, planned_v1)
    ]

    df["% Final Delivery"] = [
        safe_div(a, cp) for a, cp in zip(actual, craft_planned)
    ]

    df["Total KPI Achieved"] = df["% Final Delivery"]

    df["Deviation % v1 & CRAFT Plan"] = [
        safe_diff_div(cp, p, p)
        for cp, p in zip(craft_planned, planned_v1)
    ]

    df["Deviation % Platform & CRAFT Delivery"] = [
        safe_diff_div(a, cr, cr)
        for a, cr in zip(actual, craft_reported)
    ]


st.success("✅ Phase 7 completed — calculations added safely for all tabs")


















# =====================================================
# PHASE 8 — FINAL EXCEL OUTPUT + FORMATTING
# =====================================================

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO


FINAL_COLUMNS = [
    "Genres","Site","Unique Key","Objective/Targeting","Property/Inventory",
    "Ad Unit","Cost Format","Campaign Days","Monitoring Days",
    "Start Date","End Date","Live Date",
    "Planned Delivery v1","CRAFT Planned Delivery",
    "Actual Delivered Reporting SA","CRAFT Reported Delivery",
    "% v1 Delivery","% Final Delivery","Total KPI Achieved",
    "Deviation % v1 & CRAFT Plan","Deviation % Platform & CRAFT Delivery",
]


# ---------- Add campaign & monitoring days ----------

for df in media_plan_items.values():
    

    df["Start Date"] = campaign_start
    df["End Date"] = campaign_end

    df["Campaign Days"] = (campaign_end - campaign_start).days + 1
    df["Monitoring Days"] = (campaign_end - campaign_start).days + 1

    # df["Monitoring Days"] = df["Live Date"].apply(
    #     lambda x: (campaign_end - x).days + 1 if pd.notna(x) else None
    # )


# ---------- Write Excel ----------

output = BytesIO()

with pd.ExcelWriter(output, engine="openpyxl") as writer:

    master_frames = []

    for ad_type, df in media_plan_items.items():

        df_final = df.reindex(columns=FINAL_COLUMNS)

        # df_final.to_excel(writer, sheet_name=ad_type[:31], index=False)
        sheet_name = ad_type[:31]

        df_final.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=7   # pushes table down
        )

        ws = writer.sheets[sheet_name]


        # ---- Top info block ----

        
        ws["A1"] = "Client"
        ws["B1"] = "Samsung India"

        ws["A2"] = "Brand"
        ws["B2"] = brand_name

        ws["A3"] = "Month"
        ws["B3"] = campaign_start.strftime("%b %Y")

        ws["A4"] = "Monitoring Period"
        ws["B4"] = f"{campaign_start.strftime('%d %b')} – {campaign_end.strftime('%d %b')}"

        ws["A5"] = "Campaign Duration"
        ws["B5"] = f"{campaign_start.strftime('%d %b')} – {campaign_end.strftime('%d %b')}"


        from openpyxl.styles import PatternFill, Font

        HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
        BOLD = Font(bold=True)

        for row in range(1, 6):
            ws[f"A{row}"].font = BOLD
            ws[f"A{row}"].fill = HEADER_FILL
            ws[f"B{row}"].fill = HEADER_FILL

        master_frames.append(df_final)

    if master_frames:
        pd.concat(master_frames).to_excel(
            writer, sheet_name="Master", index=False
        )


# ---------- Formatting ----------
output.seek(0)
wb = load_workbook(output, data_only=True)

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DARK_BLUE = PatternFill("solid", fgColor="FF203764")
LIGHT_BLUE = PatternFill("solid", fgColor="FF4472C4")
GREY = PatternFill("solid", fgColor="FF808080")
GREEN = PatternFill("solid", fgColor="FF70AD47")
LIGHT_GREY = PatternFill("solid", fgColor="FF808080")

HEADER_FONT = Font(size=12,  color="FFFFFF")
BODY_FONT = Font(size=12)
# bold=True,
DETAIL_LABEL = Font(size=10, color="FFFFFF")
DETAIL_VALUE = Font(size=10, color="FFFFFF")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


BLUE_COLS = [
"Genres","Site","Unique Key","Objective/Targeting","Property/Inventory",
"Ad Unit","Cost Format","Campaign Days","Monitoring Days",
"Start Date","End Date","Live Date"
]

LIGHT_BLUE_COL = ["Deviation % v1 & CRAFT Plan"]

GREY_COLS = [
"Planned Delivery v1",
"Actual Delivered Reporting SA",
"% v1 Delivery",
"% Final Delivery",
"Total KPI Achieved"
]

GREEN_COLS = [
"CRAFT Planned Delivery",
"CRAFT Reported Delivery",
"Deviation % Platform & CRAFT Delivery"
]


for ws in wb.worksheets:

    # ===== Header details section =====
    if ws.title != "Master":
        for r in range(1,6):
            ws[f"A{r}"].font = DETAIL_LABEL
            ws[f"B{r}"].font = DETAIL_VALUE
            ws[f"A{r}"].fill = LIGHT_GREY
            ws[f"B{r}"].fill = LIGHT_GREY
            ws[f"A{r}"].border = BORDER
            ws[f"B{r}"].border = BORDER

    header_row = 1 if ws.title == "Master" else 8
    data_row = header_row + 1

    headers = {
        ws.cell(header_row, c).value: c
        for c in range(1, ws.max_column + 1)
    }

    # ===== Format header row =====
    for name, col in headers.items():

        cell = ws.cell(header_row, col)

        if name in BLUE_COLS:
            cell.fill = DARK_BLUE

        elif name in LIGHT_BLUE_COL:
            cell.fill = LIGHT_BLUE

        elif name in GREY_COLS:
            cell.fill = GREY

        elif name in GREEN_COLS:
            cell.fill = GREEN

        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # ===== Table body =====
    for r in range(data_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER

FILL_UNDER = PatternFill("solid", fgColor="FFC000")
FILL_OVER = PatternFill("solid", fgColor="FFCCFF")
FILL_NA = PatternFill("solid", fgColor="9933FF")

FONT_GREEN = Font(color="00B050")
FONT_RED = Font(color="FF0000")


def get_threshold(sheet):
    return 0.30 if "dmp" in sheet.lower() else 0.10

for ws in wb.worksheets:

    # ---------- MASTER SHEET ----------
    if ws.title == "Master":

        TABLE_HEADER_ROW = 1
        DATA_START_ROW = 2

    # ---------- AD TABS ----------
    else:

        TABLE_HEADER_ROW = 8
        DATA_START_ROW = TABLE_HEADER_ROW + 1

    headers = {cell.value: i+1 for i, cell in enumerate(ws[TABLE_HEADER_ROW])}
    WHOLE_PERCENT = [
        "% v1 Delivery",
        "% Final Delivery",
        "Total KPI Achieved",
        "Deviation % v1 & CRAFT Plan"
    ]

    TWO_DECIMAL_PERCENT = [
        "Deviation % Platform & CRAFT Delivery"
    ]

    # -------- Whole % columns --------
    for col in WHOLE_PERCENT:
        if col in headers:
            col_index = headers[col]
            # for r in range(2, ws.max_row + 1):
            DATA_START_ROW = TABLE_HEADER_ROW + 1

            for r in range(DATA_START_ROW, ws.max_row + 1):
                cell = ws.cell(r, col_index)
                if cell.value is not None:
                    cell.value = float(cell.value)
                    cell.number_format = "0%"

    # -------- 2 Decimal % column --------
    for col in TWO_DECIMAL_PERCENT:
        if col in headers:
            col_index = headers[col]
            # for r in range(2, ws.max_row + 1):
            DATA_START_ROW = TABLE_HEADER_ROW + 1

            for r in range(DATA_START_ROW, ws.max_row + 1):
                cell = ws.cell(r, col_index)
                if cell.value is not None:
                    cell.value = float(cell.value)
                    cell.number_format = "0.00%"


    col_final = headers["% Final Delivery"]
    col_kpi = headers["Total KPI Achieved"]
    col_dev = headers["Deviation % Platform & CRAFT Delivery"]

    threshold = get_threshold(ws.title)

    # for r in range(2, ws.max_row + 1):
    DATA_START_ROW = TABLE_HEADER_ROW + 1

    for r in range(DATA_START_ROW, ws.max_row + 1):

        val = ws.cell(r, col_final).value

        if val is None:
            ws.cell(r, col_final).fill = FILL_NA
        elif val < (1 - threshold):
            ws.cell(r, col_final).fill = FILL_UNDER
        elif val > (1 + threshold):
            ws.cell(r, col_final).fill = FILL_OVER

        ws.cell(
            r, col_kpi
        ).font = FONT_GREEN if val and val >= (1 - threshold) else FONT_RED

        dev = ws.cell(r, col_dev).value
        if dev is not None and abs(dev) > 0.02:
            ws.cell(r, col_dev).font = FONT_RED

    # Date formatting
    for col_name in ["Start Date", "End Date", "Live Date"]:
        c = headers[col_name]
        for r in range(DATA_START_ROW, ws.max_row + 1):
            ws.cell(r, c).number_format = "DD-MMM-YY"
    from openpyxl.utils import get_column_letter
    

    # ================= TOTAL ROW (AD TABS ONLY) =================
    # TOTAL FOR ALL SHEETS (remove condition)
    last_data_row = ws.max_row

    # 🔥 If Master — ignore blank rows
    while last_data_row > DATA_START_ROW and ws.cell(last_data_row, 1).value in [None, ""]:
        last_data_row -= 1
    total_row = last_data_row + 1

    ws.cell(total_row, 1).value = "Total"
    ws.cell(total_row, 1).fill = DARK_BLUE
    ws.cell(total_row, 1).font = HEADER_FONT
    ws.cell(total_row, 1).alignment = CENTER
    ws.cell(total_row, 1).border = BORDER

    SUM_COLS = [
            "Planned Delivery v1",
            "CRAFT Planned Delivery",
            "Actual Delivered Reporting SA",
            "CRAFT Reported Delivery",
        ]

    for col_name in SUM_COLS:
            if col_name in headers:
                col = headers[col_name]
                letter = get_column_letter(col)

                ws.cell(
                    total_row, col
                ).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"

    # ===== % TOTALS AS AVERAGE (NOT SUM) =====

    # ================= TOTAL % & DEVIATIONS — FROM TOTAL NUMBERS =================

    col_plan = headers["Planned Delivery v1"]
    col_craft_plan = headers["CRAFT Planned Delivery"]
    col_actual = headers["Actual Delivered Reporting SA"]
    col_craft_rep = headers["CRAFT Reported Delivery"]

    col_pct_v1 = headers["% v1 Delivery"]
    col_pct_final = headers["% Final Delivery"]
    col_kpi = headers["Total KPI Achieved"]
    col_dev_plan = headers["Deviation % v1 & CRAFT Plan"]
    col_dev_platform = headers["Deviation % Platform & CRAFT Delivery"]

    L_plan = get_column_letter(col_plan)
    L_craft_plan = get_column_letter(col_craft_plan)
    L_actual = get_column_letter(col_actual)
    L_craft_rep = get_column_letter(col_craft_rep)

    # % v1 Delivery
    ws.cell(total_row, col_pct_v1).value = f"={L_actual}{total_row}/{L_plan}{total_row}"

    # % Final Delivery
    ws.cell(total_row, col_pct_final).value = f"={L_actual}{total_row}/{L_craft_plan}{total_row}"

    # KPI Achieved
    ws.cell(total_row, col_kpi).value = f"={get_column_letter(col_pct_final)}{total_row}"

    # Deviation v1 vs CRAFT plan
    ws.cell(
        total_row, col_dev_plan
    ).value = f"=({L_craft_plan}{total_row}-{L_plan}{total_row})/{L_plan}{total_row}"

    # Deviation Platform vs CRAFT reported
    ws.cell(
        total_row, col_dev_platform
    ).value = f"=({L_actual}{total_row}-{L_craft_rep}{total_row})/{L_craft_rep}{total_row}"

    # formatting
    ws.cell(total_row, col_pct_v1).number_format = "0%"
    ws.cell(total_row, col_pct_final).number_format = "0%"
    ws.cell(total_row, col_kpi).number_format = "0%"
    ws.cell(total_row, col_dev_plan).number_format = "0%"
    ws.cell(total_row, col_dev_platform).number_format = "0.00%"
    
# ✅ FILL TOTAL ROW ONLY ACROSS TABLE (NOT FULL SHEET)

    last_table_col = ws.max_column   # this is your table width

    for c in range(1, last_table_col + 1):
        cell = ws.cell(total_row, c)
        cell.fill = DARK_BLUE
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    
    # PERCENT_AVG_COLS = [
    #     "% v1 Delivery",
    #     "% Final Delivery",
    #     "Total KPI Achieved",
    #     "Deviation % v1 & CRAFT Plan",
    #     "Deviation % Platform & CRAFT Delivery"
    # ]
    

    # for col_name in PERCENT_AVG_COLS:
    #     if col_name in headers:
    #         col = headers[col_name]
    #         letter = get_column_letter(col)

    #         ws.cell(
    #             total_row, col
    #         ).value = f"=AVERAGE({letter}{DATA_START_ROW}:{letter}{last_data_row})"

    #         # keep formatting same as column
    #         if col_name == "Deviation % Platform & CRAFT Delivery":
    #             ws.cell(total_row, col).number_format = "0.00%"
    #         else:
    #             ws.cell(total_row, col).number_format = "0%"

    #         # full row styling
    #     for c in range(1, ws.max_column + 1):
    #             cell = ws.cell(total_row, c)
    #             cell.fill = DARK_BLUE
    #             cell.font = HEADER_FONT
    #             cell.alignment = CENTER
    #             cell.border = BORDER

        # ================= SA REMARKS =================

    # if ws.title != "Master":

    #     remarks_row = total_row + 2

    # ws.merge_cells(
    #         start_row=remarks_row,
    #         start_column=1,
    #         end_row=remarks_row,
    #         end_column=ws.max_column
    #     )

    # ws.cell(remarks_row, 1).value = (
    #         "SA Remarks: KPI performance reviewed. Under-delivered line items will be optimised in the next phase."
    #     )

    # ws.cell(remarks_row, 1).font = Font(size=10, italic=True)
    # ws.cell(remarks_row, 1).alignment = Alignment(horizontal="left", vertical="center")


    # ================= SA REMARKS (DYNAMIC) =================
# ================= SA REMARKS (USING SAME THRESHOLD LOGIC) =================

    if ws.title != "Master":

        col_final = headers["% Final Delivery"]
        col_uk = headers["Unique Key"]

        under = []
        over = []

        threshold = get_threshold(ws.title)

        for r in range(DATA_START_ROW, last_data_row + 1):

            val = ws.cell(r, col_final).value
            uk = ws.cell(r, col_uk).value

            if val is None:
                continue

            # ✅ SAME LOGIC AS COLORING
            if val < (1 - threshold):
                under.append(str(uk))

            elif val > (1 + threshold):
                over.append(str(uk))

        remarks = ["SA Remarks:"]

        if under:
            remarks.append(
                "KPI under delivered for unique key" +
                ("s " if len(under) > 1 else " ") +
                ", ".join(under)
            )

        if over:
            remarks.append(
                "KPI over delivered for unique key" +
                ("s " if len(over) > 1 else " ") +
                " and ".join(over)
            )

        remarks_row = total_row + 2

        ws.merge_cells(
            start_row=remarks_row,
            start_column=1,
            end_row=remarks_row,
            end_column=ws.max_column
        )

        ws.cell(remarks_row, 1).value = "\n".join(remarks)

        ws.cell(remarks_row, 1).font = Font(size=10, italic=True)
        ws.cell(remarks_row, 1).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

# from openpyxl.utils import get_column_letter

# # ================= TOTAL ROW =================
# if ws.title != "Master":

#     last_data_row = ws.max_row
#     total_row = last_data_row + 1

#     ws.cell(total_row, 1).value = "Total"
#     ws.cell(total_row, 1).fill = DARK_BLUE
#     ws.cell(total_row, 1).font = HEADER_FONT
#     ws.cell(total_row, 1).alignment = CENTER
#     ws.cell(total_row, 1).border = BORDER

#     SUM_COLS = [
#         "Planned Delivery v1",
#         "CRAFT Planned Delivery",
#         "Actual Delivered Reporting SA",
#         "CRAFT Reported Delivery",
#     ]

#     for col_name in SUM_COLS:
#         if col_name in headers:
#             col = headers[col_name]
#             letter = get_column_letter(col)

#             ws.cell(total_row, col).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
#             ws.cell(total_row, col).fill = DARK_BLUE
#             ws.cell(total_row, col).font = HEADER_FONT
#             ws.cell(total_row, col).alignment = CENTER
#             ws.cell(total_row, col).border = BORDER

#     # fill full total row color
#     for c in range(1, ws.max_column + 1):
#         ws.cell(total_row, c).fill = DARK_BLUE
#         ws.cell(total_row, c).font = HEADER_FONT
#         ws.cell(total_row, c).alignment = CENTER
#         ws.cell(total_row, c).border = BORDER

# from openpyxl.utils import get_column_letter

# # ================= TOTAL ROW (ALL SHEETS) =================

# last_data_row = ws.max_row
# total_row = last_data_row + 1

# ws.cell(total_row, 1).value = "Total"
# ws.cell(total_row, 1).fill = DARK_BLUE
# ws.cell(total_row, 1).font = HEADER_FONT
# ws.cell(total_row, 1).alignment = CENTER
# ws.cell(total_row, 1).border = BORDER

# SUM_COLS = [
#     "Planned Delivery v1",
#     "CRAFT Planned Delivery",
#     "Actual Delivered Reporting SA",
#     "CRAFT Reported Delivery",
# ]

# for col_name in SUM_COLS:
#     if col_name in headers:
#         col = headers[col_name]
#         letter = get_column_letter(col)

#         ws.cell(total_row, col).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
#         ws.cell(total_row, col).fill = DARK_BLUE
#         ws.cell(total_row, col).font = HEADER_FONT
#         ws.cell(total_row, col).alignment = CENTER
#         ws.cell(total_row, col).border = BORDER

# for c in range(1, ws.max_column + 1):
#     ws.cell(total_row, c).fill = DARK_BLUE
#     ws.cell(total_row, c).font = HEADER_FONT
#     ws.cell(total_row, c).alignment = CENTER
#     ws.cell(total_row, c).border = BORDER


# # ================= SA REMARKS (AD TABS ONLY) =================

# if ws.title != "Master":

#     remarks_row = total_row + 2

#     ws.merge_cells(
#         start_row=remarks_row,
#         start_column=1,
#         end_row=remarks_row,
#         end_column=ws.max_column
#     )

#     ws.cell(remarks_row, 1).value = (
#         "SA Remarks: KPI performance reviewed. Under-delivered line items will be optimised in the next phase."
#     )

#     ws.cell(remarks_row, 1).font = Font(size=10, italic=True)
#     ws.cell(remarks_row, 1).alignment = Alignment(horizontal="left", vertical="center")

# ---------- Save back ----------

final_output = BytesIO()
wb.save(final_output)


st.success("✅ SA Report Generated Successfully")

st.download_button(
    "⬇️ Download SA Report",
    data=final_output.getvalue(),
    file_name=f"SA_Report_{brand_name}_{qt_number}.xlsx",
    # file_name="SA_Report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)




